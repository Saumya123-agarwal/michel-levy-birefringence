"""
Step 1: Prepare unified dataset from raw POM images + Excel birefringence data.

Creates: data/dataset_unified.csv with columns:
  image_path, temperature_C, thickness_um, order, material, delta_n

Usage:
  python src/prepare_dataset.py --data-root dataset/dataset_physics_SOP_liquid_crystal
"""

import argparse
import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def parse_p100_7cb(data_root: str, thickness_um: float = 9.415, order: int = 1):
    """Parse P100-7CB+ dataset: images named by temperature, Excel has delta_n."""
    folder = os.path.join(data_root, "P100-7CB+ Str +t3318")
    excel_path = os.path.join(folder, "data_bifringence.xlsx")

    wb = load_workbook(excel_path)
    ws = wb["Sheet1"]

    temp_to_dn = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        temp_c, temp_k, dn = row[0], row[1], row[2]
        if temp_c is not None and dn is not None:
            temp_to_dn[float(temp_c)] = float(dn)

    records = []
    for fname in sorted(os.listdir(folder)):
        if not fname.lower().endswith(".jpg"):
            continue
        stem = fname.replace(".jpg", "").replace(".JPG", "")
        try:
            temp = float(stem)
        except ValueError:
            continue

        best_match = None
        best_diff = float("inf")
        for t in temp_to_dn:
            diff = abs(t - temp)
            if diff < best_diff:
                best_diff = diff
                best_match = t

        if best_match is not None and best_diff <= 0.3:
            dn = temp_to_dn[best_match]
        else:
            print(f"  [WARN] No delta_n for {fname} (T={temp} C), skipping")
            continue

        records.append({
            "image_path": os.path.join(folder, fname),
            "temperature_C": temp,
            "thickness_um": thickness_um,
            "order": order,
            "material": "P100-7CB+",
            "delta_n": dn,
        })

    print(f"  P100-7CB+: {len(records)} images matched")
    return records


def parse_5pch(data_root: str, thickness_um: float = 10.0, order: int = 1):
    """Parse M7 - pure 5PCH dataset."""
    folder = os.path.join(data_root, "for POM software", "M7 - pure 5PCH")
    excel_path = os.path.join(data_root, "for POM software", "Book1.xlsx")

    wb = load_workbook(excel_path)
    ws = wb["Sheet1"]

    temp_to_dn = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        temp_c, dn = row[0], row[1]
        if temp_c is not None and dn is not None:
            try:
                dn_val = float(dn)
                temp_to_dn[float(temp_c)] = dn_val
            except (ValueError, TypeError):
                continue

    records = []
    for fname in sorted(os.listdir(folder)):
        if not fname.lower().endswith(".jpg"):
            continue
        stem = fname.replace(".jpg", "").replace(".JPG", "")
        stem = re.sub(r"-\d+$", "", stem)
        try:
            temp = float(stem)
        except ValueError:
            continue

        best_match = None
        best_diff = float("inf")
        for t in temp_to_dn:
            diff = abs(t - temp)
            if diff < best_diff:
                best_diff = diff
                best_match = t

        if best_match is not None and best_diff <= 0.6:
            dn = temp_to_dn[best_match]
        else:
            print(f"  [WARN] No delta_n for {fname} (T={temp} C), skipping")
            continue

        records.append({
            "image_path": os.path.join(folder, fname),
            "temperature_C": temp,
            "thickness_um": thickness_um,
            "order": order,
            "material": "5PCH",
            "delta_n": dn,
        })

    print(f"  5PCH: {len(records)} images matched")
    return records


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--data-root", type=str,
                        default="dataset/dataset_physics_SOP_liquid_crystal")
    parser.add_argument("--output", type=str, default="data/dataset_unified.csv")
    parser.add_argument("--p100-thickness", type=float, default=9.415)
    parser.add_argument("--p100-order", type=int, default=1)
    parser.add_argument("--pch-thickness", type=float, default=10.0)
    parser.add_argument("--pch-order", type=int, default=1)
    args = parser.parse_args()

    print("Parsing P100-7CB+ dataset...")
    rec1 = parse_p100_7cb(args.data_root, args.p100_thickness, args.p100_order)

    print("Parsing 5PCH dataset...")
    rec2 = parse_5pch(args.data_root, args.pch_thickness, args.pch_order)

    all_records = rec1 + rec2
    df = pd.DataFrame(all_records)

    df["retardation_nm"] = df["delta_n"] * df["thickness_um"] * 1000
    # Approximate Tc for normalization
    df["Tc_approx"] = df["material"].map({"P100-7CB+": 40.0, "5PCH": 53.7})
    df["reduced_temp"] = df["temperature_C"] / df["Tc_approx"]

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    df.to_csv(args.output, index=False)

    print(f"\n{'='*60}")
    print(f"Saved: {args.output}")
    print(f"Total samples: {len(df)}")
    print(df.groupby("material").agg({
        "temperature_C": ["min", "max", "count"],
        "delta_n": ["min", "max", "mean"],
    }).to_string())


if __name__ == "__main__":
    main()
